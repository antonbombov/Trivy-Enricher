# trivy_excel_reporter.py
import json
from pathlib import Path
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Константа со статусами для выпадающего списка
STATUS_OPTIONS = [
    "Опровергнута",
    "Устранена",
    "Устранение невозможно",
    "Приняты компенсирующие меры",
    "В работе"
]

# Константа с именами колонок
COLUMN_HEADERS = [
    "№",
    "Источник",
    "Путь",
    "Пакет",
    "Версия",
    "Идентификатор уязвимости",
    "Уровень критичности",
    "Срок устранения",  # Новая колонка
    "Статус",
    "Комментарий"
]


def generate_trivy_excel_report(enriched_trivy_path, output_dir=None):
    """
    Генерирует Excel отчет из обогащенного отчета Trivy
    БЕЗ дедупликации - каждая строка = отдельное вхождение уязвимости
    """
    try:
        # Загружаем обогащенный отчет
        with open(enriched_trivy_path, 'r', encoding='utf-8-sig') as f:
            trivy_data = json.load(f)

        # Определяем путь для сохранения
        if output_dir is None:
            from config_manager import load_config
            config = load_config()
            output_dir = Path(config.get('output_directory', Path(__file__).parent))
        else:
            output_dir = Path(output_dir)

        # Создаем директорию, если её нет
        output_dir.mkdir(parents=True, exist_ok=True)

        # Сохраняем в указанной директории
        output_path = output_dir / f"{enriched_trivy_path.stem}_report.xlsx"

        # Собираем ВСЕ вхождения уязвимостей (БЕЗ дедупликации)
        all_vulnerabilities = collect_all_vulnerabilities(trivy_data)

        # Получаем имя артефакта и текущую дату для заголовка
        artifact_name = get_artifact_name(enriched_trivy_path.name)
        current_date = datetime.now()

        # Генерируем Excel
        workbook = generate_excel_content(all_vulnerabilities, artifact_name, current_date)

        # Сохраняем файл
        workbook.save(output_path)

        return output_path

    except Exception as e:
        import traceback
        print(f"ОШИБКА генерации Excel отчета: {e}")
        print(f"Трассировка ошибки:")
        traceback.print_exc()
        return None


def collect_all_vulnerabilities(trivy_data):
    """
    Собирает ВСЕ вхождения уязвимостей БЕЗ какой-либо дедупликации
    Каждая запись в отчете Trivy = отдельная строка в Excel
    """
    all_vulns = []
    current_date = datetime.now()

    if 'Results' in trivy_data:
        for result in trivy_data['Results']:
            target = result.get('Target', 'Unknown')  # Берем Target как источник
            if 'Vulnerabilities' in result:
                for vuln in result['Vulnerabilities']:
                    if 'VulnerabilityID' not in vuln:
                        continue

                    pkg_path = vuln.get('PkgPath', 'N/A')

                    # Разделяем путь на root jar и остальной путь
                    root_jar = extract_root_jar(pkg_path)

                    # Формируем источник: Target + root jar (с новой строки)
                    source_value = target
                    if root_jar and root_jar != 'N/A':
                        source_value = f"{target}\n({root_jar})"

                    # Путь без root jar
                    path_without_root = pkg_path
                    if root_jar and pkg_path and pkg_path != 'N/A':
                        # Убираем root jar из начала пути
                        if pkg_path.startswith(root_jar + '/'):
                            path_without_root = pkg_path[len(root_jar) + 1:]
                        elif pkg_path == root_jar:
                            path_without_root = "(root)"

                        # Убираем ВСЕ упоминания BOOT-INF/ из пути (где бы они ни были)
                        if path_without_root and path_without_root != 'N/A' and path_without_root != "(root)":
                            path_without_root = path_without_root.replace('BOOT-INF/', '')

                    # Проверяем наличие эксплойтов
                    sploitscan = vuln.get('sploitscan', {})
                    has_exploits = has_any_exploits(sploitscan)

                    # Получаем severity
                    severity = vuln.get('Severity', 'UNKNOWN')

                    # Рассчитываем срок устранения
                    remediation_date = calculate_remediation_date(current_date, severity, has_exploits)

                    all_vulns.append({
                        'source': source_value,  # Источник с root jar
                        'path': path_without_root,  # Путь без root jar и BOOT-INF
                        'package': vuln.get('PkgName', 'Unknown Package'),
                        'version': vuln.get('InstalledVersion', 'Unknown'),
                        'vulnerability_id': vuln['VulnerabilityID'],
                        'severity': severity,
                        'remediation_date': remediation_date,  # Срок устранения
                        'status': '',  # Пустой для ручного заполнения
                        'comment': ''  # Пустой для ручного заполнения
                    })

    return all_vulns


def extract_root_jar(pkg_path):
    """
    Извлекает корневой JAR файл из пути
    Полная копия функции из trivy_html_reporter.py
    """
    if not pkg_path:
        return None

    parts = pkg_path.split('/')
    root = parts[0]
    if root.endswith('.jar') or root.endswith('.war') or root.endswith('.ear'):
        return root

    return None


def has_any_exploits(sploitscan):
    """
    Проверяет, есть ли эксплойты в любом источнике
    Полная копия функции из trivy_html_reporter.py
    """
    # Если sploitscan - пустой список (ошибка сканирования)
    if isinstance(sploitscan, list) and len(sploitscan) == 0:
        return False

    # Если sploitscan - не словарь или пустой словарь
    if not isinstance(sploitscan, dict) or not sploitscan:
        return False

    # 1. Проверяем GitHub PoCs
    github_data = sploitscan.get('GitHub Data')
    if github_data and isinstance(github_data, dict):
        github_pocs = github_data.get('pocs', [])
        if github_pocs and len(github_pocs) > 0:
            return True

    # 2. Проверяем ExploitDB Data (по полю id)
    exploitdb_list = sploitscan.get('ExploitDB Data', [])
    if exploitdb_list:
        for item in exploitdb_list:
            if isinstance(item, dict) and item.get('id'):
                return True

    # 3. Проверяем NVD Data exploits
    nvd_data = sploitscan.get('NVD Data')
    if nvd_data and isinstance(nvd_data, dict):
        nvd_exploits = nvd_data.get('exploits', [])
        if nvd_exploits and len(nvd_exploits) > 0:
            return True

    # 4. Проверяем Metasploit Data modules
    metasploit_data = sploitscan.get('Metasploit Data')
    if metasploit_data and isinstance(metasploit_data, dict):
        metasploit_modules = metasploit_data.get('modules', [])
        if metasploit_modules:
            for module in metasploit_modules:
                if isinstance(module, dict) and module.get('url'):
                    return True

    return False


def calculate_remediation_date(base_date, severity, has_exploits):
    """
    Рассчитывает срок устранения уязвимости
    """
    # Базовый срок в месяцах в зависимости от severity
    if severity == 'CRITICAL':
        months = 6
    elif severity == 'HIGH':
        months = 9
    else:  # MEDIUM, LOW, UNKNOWN
        months = 12

    # Если есть эксплойты, сокращаем на 3 месяца
    if has_exploits:
        months = max(1, months - 3)  # Минимум 1 месяц

    # Рассчитываем дату
    remediation_date = base_date + relativedelta(months=months)

    return remediation_date.strftime('%d.%m.%Y')


def get_artifact_name(report_filename):
    """
    Извлекает имя артефакта из имени файла (без .json)
    Полная копия функции из trivy_html_reporter.py
    """
    return Path(report_filename).stem.replace('_enriched', '')


def generate_excel_content(vulnerabilities, artifact_name, current_date):
    """
    Генерирует Excel файл с данными уязвимостей
    """
    # Создаем новую рабочую книгу
    wb = openpyxl.Workbook()
    ws = wb.active

    # Устанавливаем название листа
    ws.title = "SCA Анализ"  # Новое название листа

    # Добавляем информационный блок в начало
    add_info_block(ws, artifact_name, current_date)

    # Настройка стилей
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Заголовки начинаются с 5 строки (после информационного блока)
    start_row = 5

    # Создаем заголовки
    for col_num, header in enumerate(COLUMN_HEADERS, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Заполняем данные
    for row_num, vuln in enumerate(vulnerabilities, start_row + 1):
        # №
        cell = ws.cell(row=row_num, column=1, value=row_num - start_row)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Источник
        cell = ws.cell(row=row_num, column=2, value=vuln['source'])
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Путь
        cell = ws.cell(row=row_num, column=3, value=vuln['path'])
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Пакет
        cell = ws.cell(row=row_num, column=4, value=vuln['package'])
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')

        # Версия
        cell = ws.cell(row=row_num, column=5, value=vuln['version'])
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')

        # Идентификатор уязвимости
        cell = ws.cell(row=row_num, column=6, value=vuln['vulnerability_id'])
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Уровень критичности
        severity = vuln['severity']
        cell = ws.cell(row=row_num, column=7, value=severity)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Срок устранения
        cell = ws.cell(row=row_num, column=8, value=vuln['remediation_date'])
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Статус (пустой)
        cell = ws.cell(row=row_num, column=9, value='')
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Комментарий (пустой)
        cell = ws.cell(row=row_num, column=10, value='')
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Добавляем выпадающие списки для колонки "Статус" (колонка I, индекс 9)
    if vulnerabilities:  # Добавляем только если есть данные
        add_dropdown_lists(ws, len(vulnerabilities), start_row)

    # Настраиваем ширину колонок
    set_column_widths(ws)

    # Добавляем фильтры
    ws.auto_filter.ref = f"A{start_row}:J{len(vulnerabilities) + start_row}"

    # Замораживаем строку с заголовками
    ws.freeze_panes = f'A{start_row + 1}'

    return wb


def add_info_block(worksheet, artifact_name, current_date):
    """
    Добавляет информационный блок в начало отчета
    """
    # Форматируем дату
    date_str = current_date.strftime('%d.%m.%Y')

    # Стиль для информационного блока
    info_font = Font(name='Arial', size=11, bold=True)
    info_alignment = Alignment(horizontal='left', vertical='center')

    # Добавляем наименование проверяемого объекта
    cell = worksheet.cell(row=1, column=1, value=f"Объект проверки: {artifact_name}")
    cell.font = info_font
    cell.alignment = info_alignment

    # Добавляем дату построения отчета
    cell = worksheet.cell(row=2, column=1, value=f"Дата построения отчета: {date_str}")
    cell.font = info_font
    cell.alignment = info_alignment

    # Добавляем формулировку о необходимости обновления
    cell = worksheet.cell(row=3, column=1,
                          value="Необходимо обновить уязвимые компоненты до актуальных версий в указанный срок или обосновать отсутствие такой возможности")
    cell.font = info_font
    cell.alignment = info_alignment

    # Объединяем ячейки для информационного блока (A1:J1, A2:J2, A3:J3)
    for row in range(1, 4):
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMN_HEADERS))

    # Добавляем пустую строку перед заголовками
    worksheet.row_dimensions[4].height = 10


def add_dropdown_lists(worksheet, num_rows, start_row):
    """
    Добавляет выпадающие списки со статусами в колонку I (Статус)
    Запрещает ввод произвольного текста - только выбор из списка
    """
    if num_rows == 0:
        return

    # Создаем отдельный лист со справочником статусов
    wb = worksheet.parent
    hidden_sheet = wb.create_sheet("_status_reference")
    hidden_sheet.sheet_state = 'hidden'  # Скрываем лист

    # Записываем статусы в скрытый лист
    for i, status in enumerate(STATUS_OPTIONS, 1):
        hidden_sheet.cell(row=i, column=1, value=status)

    # Создаем именованный диапазон для статусов
    ref_range = f"'_status_reference'!$A$1:$A${len(STATUS_OPTIONS)}"

    # Добавляем выпадающие списки для каждой строки с данными
    for row_num in range(start_row + 1, start_row + num_rows + 1):
        cell_ref = f"I{row_num}"  # Статус теперь в колонке I (индекс 9)

        # Создаем правило валидации данных
        dv = openpyxl.worksheet.datavalidation.DataValidation(
            type='list',
            formula1=ref_range,
            allow_blank=True,  # Разрешаем пустое значение
            showErrorMessage=True,  # Показывать сообщение об ошибке
            showInputMessage=True,  # Показывать подсказку при выборе ячейки
            promptTitle='Выберите статус',
            prompt='Пожалуйста, выберите статус из выпадающего списка',
            errorTitle='Недопустимое значение',
            error='Вы можете выбрать ТОЛЬКО значение из выпадающего списка! Ввод произвольного текста запрещен.'
        )

        worksheet.add_data_validation(dv)
        dv.add(cell_ref)


def set_column_widths(worksheet):
    """
    Устанавливает оптимальную ширину колонок
    """
    # №
    worksheet.column_dimensions['A'].width = 8

    # Источник
    worksheet.column_dimensions['B'].width = 40  # Увеличена для размещения root jar

    # Путь
    worksheet.column_dimensions['C'].width = 50

    # Пакет
    worksheet.column_dimensions['D'].width = 25

    # Версия
    worksheet.column_dimensions['E'].width = 20

    # Идентификатор уязвимости
    worksheet.column_dimensions['F'].width = 20

    # Уровень критичности
    worksheet.column_dimensions['G'].width = 15

    # Срок устранения (новая колонка)
    worksheet.column_dimensions['H'].width = 15

    # Статус
    worksheet.column_dimensions['I'].width = 25

    # Комментарий
    worksheet.column_dimensions['J'].width = 40


def main():
    """
    Основная функция для тестирования
    """
    script_dir = Path(__file__).parent
    enriched_files = list(script_dir.glob("*_enriched.json"))

    if not enriched_files:
        print("Нет обогащенных отчетов Trivy")
        return

    for enriched_file in enriched_files:
        print(f"Генерация Excel отчета для: {enriched_file.name}")
        generate_trivy_excel_report(enriched_file)


if __name__ == "__main__":
    main()