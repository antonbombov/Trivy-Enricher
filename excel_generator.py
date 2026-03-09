import openpyxl
import sys
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ptai_parser import PTAIParser


def main():
    if len(sys.argv) < 2:
        print("Использование: python excel_generator.py <путь_к_html_файлу>")
        print("Пример: python excel_generator.py report.html")
        sys.exit(1)

    html_file = sys.argv[1]

    try:
        with open(html_file, 'r', encoding='utf-8') as f:
            html = f.read()
    except FileNotFoundError:
        print(f"Ошибка: Файл '{html_file}' не найден")
        sys.exit(1)
    except Exception as e:
        print(f"Ошибка при чтении файла: {e}")
        sys.exit(1)

    parser = PTAIParser(html)
    print(f"Проект: {parser.project_name}")
    print(f"Найдено уязвимостей: {len(parser.vulnerabilities)}")

    if not parser.vulnerabilities:
        print("Уязвимости не найдены")
        return

    data = parser.get_data_for_excel()

    # Создаем Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = parser.project_name[:31]

    # Заголовки
    headers = ['№', 'ID уязвимости', 'Тип уязвимости', 'Класс и метод / Уязвимый файл',
               'Комментарий', 'Статус', 'CWSS / Vисп', 'Компенсирующие меры']

    # Стиль границ для всех ячеек
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Данные
    for row, item in enumerate(data, 2):
        # Номер строки
        cell_num = ws.cell(row=row, column=1, value=row - 1)
        cell_num.border = thin_border
        cell_num.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Остальные колонки
        for col, key in enumerate(['ID уязвимости', 'Тип уязвимости', 'Класс и метод / Уязвимый файл',
                                   'Комментарий', 'Статус', 'CWSS / Vисп', 'Компенсирующие меры'], 2):
            value = item[key]

            # Если статус "Опровергнута" и это колонки CWSS или Компенсирующие меры
            if item['Статус'].lower() == 'опровергнута' and col in [7, 8]:
                value = '—'  # Длинный прочерк

            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border

            # Выравнивание
            if col in [1, 2, 6, 7, 8]:  # №, ID, Статус, CWSS, Компенсирующие меры
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Ширина колонок
    ws.column_dimensions['A'].width = 5  # №
    ws.column_dimensions['B'].width = 15  # ID
    ws.column_dimensions['C'].width = 40  # Тип
    ws.column_dimensions['D'].width = 50  # Файл
    ws.column_dimensions['E'].width = 40  # Комментарий
    ws.column_dimensions['F'].width = 15  # Статус
    ws.column_dimensions['G'].width = 15  # CWSS
    ws.column_dimensions['H'].width = 30  # Компенсирующие меры

    base_name = os.path.splitext(os.path.basename(html_file))[0]
    output_file = f"{base_name}_report.xlsx"

    wb.save(output_file)
    print(f"Отчет сохранен: {output_file}")


if __name__ == "__main__":
    main()