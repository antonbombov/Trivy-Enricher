# excel_reporter.py
import json
from pathlib import Path
from datetime import datetime
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ptai_processor import prepare_ptai_excel_data, PTAIParser
from excel_styles import (
    HEADER_FONT, HEADER_FILL, CELL_BORDER, INFO_FONT,
    apply_header_style, apply_cell_style, SCA_COLUMN_WIDTHS, PTAI_COLUMN_WIDTHS
)

# Константа со статусами для выпадающего списка
STATUS_OPTIONS = [
    "Опровергнута",
    "Устранена",
    "Устранение невозможно",
    "Приняты компенсирующие меры",
    "В работе"
]

# Константа с именами колонок для SCA анализа
SCA_COLUMN_HEADERS = [
    "№",
    "Источник",
    "Путь",
    "Пакет",
    "Версия",
    "Идентификатор уязвимости",
    "Уровень критичности",
    "Срок устранения",
    "Статус",
    "Комментарий"
]

# Константа с именами колонок для PTAI анализа
PTAI_COLUMN_HEADERS = [
    '№',
    'ID уязвимости',
    'Тип уязвимости',
    'Класс и метод / Уязвимый файл',
    'Комментарий',
    'Статус',
    'CWSS / Vисп',
    'Компенсирующие меры'
]


def generate_excel_report(enriched_trivy_path, output_dir, ptai_html_path=None):
    """
    Основной метод генерации Excel отчета
    Создает Excel файл с листами SCA Анализ и (опционально) PTAI Анализ
    """
    try:
        # Создаем новую рабочую книгу
        wb = openpyxl.Workbook()

        # Удаляем дефолтный лист
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Добавляем лист с SCA анализом
        print("   Добавление листа SCA Анализ...")
        add_sca_sheet(wb, enriched_trivy_path)

        # Добавляем лист с PTAI анализом, только если файл существует и имена совпадают
        if ptai_html_path and Path(ptai_html_path).exists():
            # Проверяем соответствие имен (для надежности)
            trivy_name = Path(enriched_trivy_path).stem.replace('_enriched', '')
            ptai_name = Path(ptai_html_path).stem

            if trivy_name == ptai_name:
                print(f"   Добавление листа PTAI Анализ из: {Path(ptai_html_path).name}...")
                add_ptai_sheet(wb, ptai_html_path)
            else:
                print(f"   ⚠️ Пропуск PTAI листа: имена файлов не совпадают")
                print(f"      Ожидалось: {trivy_name}.html, получено: {ptai_name}.html")
        else:
            if ptai_html_path is None:
                print("   ⚠️ PTAI отчет не найден, создается только лист SCA Анализ")
            else:
                print(f"   ⚠️ PTAI отчет не существует: {ptai_html_path}")

        # Определяем путь для сохранения
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # Формируем имя выходного файла
        base_name = Path(enriched_trivy_path).stem.replace('_enriched', '')
        output_path = output_dir / f"{base_name}_report.xlsx"

        # Сохраняем файл
        wb.save(output_path)
        print(f"   ✅ Excel файл сохранен: {output_path}")

        # Информация о созданных листах
        sheets_created = ["SCA Анализ"]
        if ptai_html_path and Path(ptai_html_path).exists() and trivy_name == ptai_name:
            sheets_created.append("PTAI Анализ")
        print(f"   📊 Созданы листы: {', '.join(sheets_created)}")

        return output_path

    except Exception as e:
        import traceback
        print(f"❌ ОШИБКА генерации Excel отчета: {e}")
        print(f"   Трассировка ошибки:")
        traceback.print_exc()
        return None


def group_vulnerabilities_by_artifact(vulnerabilities):
    """
    Группирует уязвимости по уникальному артефакту (источник + путь + пакет + версия)
    Объединяет все CVE, найденные для одного и того же артефакта
    """
    from collections import defaultdict

    groups = defaultdict(lambda: {
        'vulnerability_ids': set(),
        'severities': [],
        'severity_levels': [],
        'remediation_dates': [],
        'statuses': set(),
        'comments': set(),
        'source': None,
        'path': None,
        'package': None,
        'version': None
    })

    severity_order = {'CRITICAL': 4, 'HIGH': 3, 'MEDIUM': 2, 'LOW': 1, 'UNKNOWN': 0}

    for vuln in vulnerabilities:
        # Ключ группировки: источник + путь + пакет + версия
        key = (vuln['source'], vuln['path'], vuln['package'], vuln['version'])
        group = groups[key]

        # Сохраняем базовую информацию
        group['source'] = vuln['source']
        group['path'] = vuln['path']
        group['package'] = vuln['package']
        group['version'] = vuln['version']

        # ID уязвимостей - используем set для уникальности
        group['vulnerability_ids'].add(vuln['vulnerability_id'])

        # Уровни критичности
        severity = vuln['severity']
        group['severities'].append(severity)
        group['severity_levels'].append((severity_order.get(severity, 0), severity))

        # Сроки устранения с их приоритетом
        try:
            rem_date = datetime.strptime(vuln['remediation_date'], '%d.%m.%Y')
            group['remediation_dates'].append((severity_order.get(severity, 0), rem_date, vuln['remediation_date']))
        except:
            group['remediation_dates'].append((severity_order.get(severity, 0), None, vuln['remediation_date']))

        # Статус и комментарий
        if vuln.get('status'):
            group['statuses'].add(vuln['status'])
        if vuln.get('comment'):
            group['comments'].add(vuln['comment'])

    # Формируем результат
    result = []
    for key, group in groups.items():
        # Объединяем идентификаторы уязвимостей (уникальные, сортированные)
        vuln_ids_str = '\n'.join(sorted(group['vulnerability_ids']))

        # Объединяем уровни критичности (уникальные, сортированные по важности)
        unique_severities = []
        seen = set()
        for _, sev in sorted(group['severity_levels'], key=lambda x: x[0], reverse=True):
            if sev not in seen:
                seen.add(sev)
                unique_severities.append(sev)
        severity_str = '\n'.join(unique_severities)

        # Выбираем срок устранения: берем дату для наивысшего уровня критичности
        remediation_date = ''
        if group['remediation_dates']:
            sorted_dates = sorted(group['remediation_dates'], key=lambda x: x[0], reverse=True)
            remediation_date = sorted_dates[0][2]

        # Статус и комментарий
        status_str = '\n'.join(sorted(group['statuses'])) if group['statuses'] else ''
        comment_str = '\n'.join(sorted(group['comments'])) if group['comments'] else ''

        result.append({
            'source': group['source'],
            'path': group['path'],
            'package': group['package'],
            'version': group['version'],
            'vulnerability_id': vuln_ids_str,
            'severity': severity_str,
            'remediation_date': remediation_date,
            'status': status_str,
            'comment': comment_str
        })

    return result


def add_sca_sheet(workbook, enriched_trivy_path):
    """
    Добавляет лист с SCA анализом из Trivy с группировкой по уникальному пути
    """
    # Загружаем обогащенный отчет
    with open(enriched_trivy_path, 'r', encoding='utf-8-sig') as f:
        trivy_data = json.load(f)

    # Создаем лист
    ws = workbook.create_sheet("SCA Анализ", 0)  # Вставляем первым

    # Собираем ВСЕ вхождения уязвимостей
    all_vulnerabilities = collect_all_vulnerabilities(trivy_data)

    # Группируем по уникальному пути (источник + путь)
    grouped_vulnerabilities = group_vulnerabilities_by_artifact(all_vulnerabilities)

    # Получаем имя артефакта и текущую дату
    artifact_name = get_artifact_name(Path(enriched_trivy_path).name)
    current_date = datetime.now()

    # Добавляем информационный блок
    add_info_block(ws, artifact_name, current_date)

    # Создаем заголовки
    start_row = 5
    for col_num, header in enumerate(SCA_COLUMN_HEADERS, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = CELL_BORDER

    # Заполняем данные (уже сгруппированные)
    for row_num, vuln in enumerate(grouped_vulnerabilities, start_row + 1):
        # №
        cell = ws.cell(row=row_num, column=1, value=row_num - start_row)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Источник
        cell = ws.cell(row=row_num, column=2, value=vuln['source'])
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Путь (объединенный)
        cell = ws.cell(row=row_num, column=3, value=vuln['path'])
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Пакет
        cell = ws.cell(row=row_num, column=4, value=vuln['package'])
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center')

        # Версия
        cell = ws.cell(row=row_num, column=5, value=vuln['version'])
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center')

        # Идентификатор уязвимости (объединенный)
        cell = ws.cell(row=row_num, column=6, value=vuln['vulnerability_id'])
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Уровень критичности (объединенный) - с переносом по словам
        cell = ws.cell(row=row_num, column=7, value=vuln['severity'])
        cell.border = CELL_BORDER
        # Принудительно устанавливаем перенос текста
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, text_rotation=0)

        # Принудительно устанавливаем высоту строки для автоматического расчета
        ws.row_dimensions[row_num].height = None

        # Срок устранения (от наивысшего уровня критичности)
        cell = ws.cell(row=row_num, column=8, value=vuln['remediation_date'])
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Статус (объединенный)
        cell = ws.cell(row=row_num, column=9, value=vuln['status'])
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Комментарий (объединенный)
        cell = ws.cell(row=row_num, column=10, value=vuln['comment'])
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Дополнительно проходим по всем строкам столбца G и принудительно применяем перенос
    for row_num in range(start_row + 1, start_row + len(grouped_vulnerabilities) + 1):
        cell = ws.cell(row=row_num, column=7)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, text_rotation=0)

    # Добавляем выпадающие списки для колонки "Статус"
    if grouped_vulnerabilities:
        add_dropdown_lists(ws, len(grouped_vulnerabilities), start_row, workbook)

    # Настраиваем ширину колонок
    set_sca_column_widths(ws)

    # Добавляем фильтры
    ws.auto_filter.ref = f"A{start_row}:J{len(grouped_vulnerabilities) + start_row}"

    # Замораживаем строку с заголовками
    ws.freeze_panes = f'A{start_row + 1}'


def add_ptai_sheet(workbook, html_file_path):
    """
    Добавляет лист с PTAI анализом в существующий workbook
    """
    # Подготавливаем данные из PTAI отчета
    data, project_name = prepare_ptai_excel_data(html_file_path)

    if not data:
        print(f"   ⚠️ Нет данных для PTAI листа")
        return

    # Создаем лист
    ws = workbook.create_sheet("PTAI Анализ")

    # Добавляем информационный блок
    current_date = datetime.now()

    # Наименование проверяемого объекта
    cell = ws.cell(row=1, column=1, value=f"Объект проверки: {project_name}")
    cell.font = INFO_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')

    # Дата построения отчета
    cell = ws.cell(row=2, column=1, value=f"Дата построения отчета: {current_date.strftime('%d.%m.%Y')}")
    cell.font = INFO_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')

    # Источник данных
    cell = ws.cell(row=3, column=1, value=f"Источник: {Path(html_file_path).name}")
    cell.font = INFO_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')

    # Пустая строка перед заголовками
    ws.row_dimensions[4].height = 10

    # Заголовки
    for col, header in enumerate(PTAI_COLUMN_HEADERS, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = CELL_BORDER

    # Данные
    for row, item in enumerate(data, 6):
        # Номер строки
        cell_num = ws.cell(row=row, column=1, value=row - 5)
        cell_num.border = CELL_BORDER
        cell_num.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # ID уязвимости
        cell = ws.cell(row=row, column=2, value=item['ID уязвимости'])
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Тип уязвимости
        cell = ws.cell(row=row, column=3, value=item['Тип уязвимости'])
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Класс и метод / Уязвимый файл
        cell = ws.cell(row=row, column=4, value=item['Класс и метод / Уязвимый файл'])
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Комментарий
        cell = ws.cell(row=row, column=5, value=item['Комментарий'])
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Статус
        status_value = item['Статус']
        cell = ws.cell(row=row, column=6, value=status_value)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # CWSS / Vисп
        cwss_value = item['CWSS / Vисп']
        # Если статус "Опровергнута", ставим прочерк
        if status_value.lower() == 'опровергнута' and not cwss_value:
            cwss_value = '—'
        cell = ws.cell(row=row, column=7, value=cwss_value)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Компенсирующие меры
        measures_value = item['Компенсирующие меры']
        # Если статус "Опровергнута", ставим прочерк
        if status_value.lower() == 'опровергнута' and not measures_value:
            measures_value = '—'
        cell = ws.cell(row=row, column=8, value=measures_value)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Настраиваем ширину колонок
    set_ptai_column_widths(ws)

    # Добавляем фильтры
    if data:
        ws.auto_filter.ref = f"A5:H{len(data) + 5}"

    # Замораживаем строку с заголовками
    ws.freeze_panes = 'A6'

    print(f"   ✅ Добавлено {len(data)} записей в лист PTAI Анализ")


def collect_all_vulnerabilities(trivy_data):
    """
    Собирает ВСЕ вхождения уязвимостей БЕЗ какой-либо дедупликации
    """
    all_vulns = []
    current_date = datetime.now()

    if 'Results' in trivy_data:
        for result in trivy_data['Results']:
            target = result.get('Target', 'Unknown')
            if 'Vulnerabilities' in result:
                for vuln in result['Vulnerabilities']:
                    if 'VulnerabilityID' not in vuln:
                        continue

                    pkg_path = vuln.get('PkgPath', 'N/A')
                    root_jar = extract_root_jar(pkg_path)

                    # Формируем источник
                    source_value = target
                    if root_jar and root_jar != 'N/A':
                        source_value = f"{target}\n({root_jar})"

                    # Путь без root jar
                    path_without_root = pkg_path
                    if root_jar and pkg_path and pkg_path != 'N/A':
                        if pkg_path.startswith(root_jar + '/'):
                            path_without_root = pkg_path[len(root_jar) + 1:]
                        elif pkg_path == root_jar:
                            path_without_root = "(root)"

                        if path_without_root and path_without_root != 'N/A' and path_without_root != "(root)":
                            path_without_root = path_without_root.replace('BOOT-INF/', '')

                    # Проверяем наличие эксплойтов
                    sploitscan = vuln.get('sploitscan', {})
                    has_exploits = has_any_exploits(sploitscan)

                    # Получаем severity
                    severity = vuln.get('Severity', 'UNKNOWN')

                    # Рассчитываем срок устранения
                    remediation_date = calculate_remediation_date(current_date, severity, has_exploits)

                    all_vulns.append({
                        'source': source_value,
                        'path': path_without_root,
                        'package': vuln.get('PkgName', 'Unknown Package'),
                        'version': vuln.get('InstalledVersion', 'Unknown'),
                        'vulnerability_id': vuln['VulnerabilityID'],
                        'severity': severity,
                        'remediation_date': remediation_date,
                        'status': '',
                        'comment': ''
                    })

    return all_vulns


def extract_root_jar(pkg_path):
    """
    Извлекает корневой JAR файл из пути
    """
    if not pkg_path or pkg_path == 'N/A':
        return None

    parts = pkg_path.split('/')
    root = parts[0]
    if root.endswith('.jar') or root.endswith('.war') or root.endswith('.ear'):
        return root

    return None


def has_any_exploits(sploitscan):
    """
    Проверяет, есть ли эксплойты в любом источнике
    """
    if isinstance(sploitscan, list) and len(sploitscan) == 0:
        return False

    if not isinstance(sploitscan, dict) or not sploitscan:
        return False

    # GitHub PoCs
    github_data = sploitscan.get('GitHub Data')
    if github_data and isinstance(github_data, dict):
        github_pocs = github_data.get('pocs', [])
        if github_pocs and len(github_pocs) > 0:
            return True

    # ExploitDB
    exploitdb_list = sploitscan.get('ExploitDB Data', [])
    if exploitdb_list:
        for item in exploitdb_list:
            if isinstance(item, dict) and item.get('id'):
                return True

    # NVD exploits
    nvd_data = sploitscan.get('NVD Data')
    if nvd_data and isinstance(nvd_data, dict):
        nvd_exploits = nvd_data.get('exploits', [])
        if nvd_exploits and len(nvd_exploits) > 0:
            return True

    # Metasploit
    metasploit_data = sploitscan.get('Metasploit Data')
    if metasploit_data and isinstance(metasploit_data, dict):
        metasploit_modules = metasploit_data.get('modules', [])
        if metasploit_modules:
            for module in metasploit_modules:
                if isinstance(module, dict) and module.get('url'):
                    return True

    return False


def calculate_remediation_date(base_date, severity, has_exploits):
    """
    Рассчитывает срок устранения уязвимости
    """
    if severity == 'CRITICAL':
        months = 6
    elif severity == 'HIGH':
        months = 9
    else:  # MEDIUM, LOW, UNKNOWN
        months = 12

    if has_exploits:
        months = max(1, months - 3)

    remediation_date = base_date + relativedelta(months=months)
    return remediation_date.strftime('%d.%m.%Y')


def get_artifact_name(report_filename):
    """
    Извлекает имя артефакта из имени файла
    """
    return Path(report_filename).stem.replace('_enriched', '')


def add_info_block(worksheet, artifact_name, current_date):
    """
    Добавляет информационный блок в начало отчета
    """
    date_str = current_date.strftime('%d.%m.%Y')

    # Объект проверки
    cell = worksheet.cell(row=1, column=1, value=f"Объект проверки: {artifact_name}")
    cell.font = INFO_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')

    # Дата построения отчета
    cell = worksheet.cell(row=2, column=1, value=f"Дата построения отчета: {date_str}")
    cell.font = INFO_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')

    # Формулировка
    cell = worksheet.cell(row=3, column=1,
                          value="Необходимо обновить уязвимые компоненты до актуальных версий в указанный срок или обосновать отсутствие такой возможности")
    cell.font = INFO_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')

    # Объединяем ячейки
    for row in range(1, 4):
        worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(SCA_COLUMN_HEADERS))

    worksheet.row_dimensions[4].height = 10


def add_dropdown_lists(worksheet, num_rows, start_row, workbook):
    """
    Добавляет выпадающие списки со статусами в колонку I (Статус)
    """
    if num_rows == 0:
        return

    # Создаем скрытый лист со справочником
    hidden_sheet = workbook.create_sheet("_status_reference")
    hidden_sheet.sheet_state = 'hidden'

    for i, status in enumerate(STATUS_OPTIONS, 1):
        hidden_sheet.cell(row=i, column=1, value=status)

    ref_range = f"'_status_reference'!$A$1:$A${len(STATUS_OPTIONS)}"

    for row_num in range(start_row + 1, start_row + num_rows + 1):
        cell_ref = f"I{row_num}"

        dv = openpyxl.worksheet.datavalidation.DataValidation(
            type='list',
            formula1=ref_range,
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
            promptTitle='Выберите статус',
            prompt='Пожалуйста, выберите статус из выпадающего списка',
            errorTitle='Недопустимое значение',
            error='Вы можете выбрать ТОЛЬКО значение из выпадающего списка!'
        )

        worksheet.add_data_validation(dv)
        dv.add(cell_ref)


def set_sca_column_widths(worksheet):
    """
    Устанавливает ширину колонок для SCA листа
    """
    for col, width in SCA_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[col].width = width


def set_ptai_column_widths(worksheet):
    """
    Устанавливает ширину колонок для PTAI листа
    """
    for col, width in PTAI_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[col].width = width


def main():
    """
    Основная функция для тестирования
    """
    script_dir = Path(__file__).parent
    enriched_files = list(script_dir.glob("*_enriched.json"))

    if not enriched_files:
        print("Нет обогащенных отчетов Trivy")
        return

    for enriched_file in enriched_files:
        print(f"\nГенерация Excel отчета для: {enriched_file.name}")

        # Для теста ищем PTAI отчет в папке PTAI рядом
        ptai_dir = script_dir / "PTAI"
        ptai_report = None
        if ptai_dir.exists():
            base_name = enriched_file.stem.replace('_enriched', '')
            potential_ptai = ptai_dir / f"{base_name}.html"
            if potential_ptai.exists():
                ptai_report = potential_ptai
                print(f"   Найден PTAI отчет: {ptai_report.name}")
            else:
                print(f"   PTAI отчет не найден для {base_name}")

        generate_excel_report(enriched_file, script_dir, ptai_report)


if __name__ == "__main__":
    main()