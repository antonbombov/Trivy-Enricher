# excel_styles.py
from openpyxl.styles import Font, PatternFill, Border, Side

# ============================================================================
# Единые стили для всех Excel отчетов
# ============================================================================

# Цвета
COLOR_HEADER_BLUE = '2F75B5'  # Синий для заголовков
COLOR_HEADER_TEXT = 'FFFFFF'  # Белый для текста заголовков
COLOR_BORDER = '000000'       # Черный для границ

# ============================================================================
# Шрифты
# ============================================================================

# Шрифт для заголовков
HEADER_FONT = Font(
    name='Arial',
    size=11,
    bold=True,
    color=COLOR_HEADER_TEXT
)

# Шрифт для информационного блока
INFO_FONT = Font(
    name='Arial',
    size=11,
    bold=True
)

# Шрифт для обычных ячеек
NORMAL_FONT = Font(
    name='Arial',
    size=10
)

# ============================================================================
# Заливка
# ============================================================================

# Заливка для заголовков
HEADER_FILL = PatternFill(
    start_color=COLOR_HEADER_BLUE,
    end_color=COLOR_HEADER_BLUE,
    fill_type='solid'
)

# ============================================================================
# Границы
# ============================================================================

# Тонкая граница для всех ячеек
THIN_BORDER = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)

# Толстая граница для выделения
THICK_BORDER = Border(
    left=Side(style='medium', color=COLOR_BORDER),
    right=Side(style='medium', color=COLOR_BORDER),
    top=Side(style='medium', color=COLOR_BORDER),
    bottom=Side(style='medium', color=COLOR_BORDER)
)

# Граница только снизу (для разделителей)
BOTTOM_BORDER = Border(
    bottom=Side(style='thin', color=COLOR_BORDER)
)

# ============================================================================
# Комбинированные стили (для удобства)
# ============================================================================

# Полный стиль для заголовков (шрифт + заливка + границы)
HEADER_STYLE = {
    'font': HEADER_FONT,
    'fill': HEADER_FILL,
    'border': THIN_BORDER
}

# Стиль для ячеек с данными
CELL_STYLE = {
    'font': NORMAL_FONT,
    'border': THIN_BORDER
}

# ============================================================================
# Цвета для уровней критичности
# ============================================================================

SEVERITY_COLORS = {
    'CRITICAL': 'FF0000',  # Красный
    'HIGH': 'FF9900',      # Оранжевый
    'MEDIUM': 'FFFF00',    # Желтый
    'LOW': '00FF00',       # Зеленый
    'UNKNOWN': '808080'    # Серый
}

# ============================================================================
# Функции для применения стилей
# ============================================================================

def apply_header_style(cell):
    """
    Применяет стиль заголовка к ячейке
    """
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER


def apply_cell_style(cell):
    """
    Применяет стиль обычной ячейки
    """
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER


def apply_severity_style(cell, severity):
    """
    Применяет стиль в зависимости от уровня критичности
    """
    cell.border = THIN_BORDER
    color = SEVERITY_COLORS.get(severity.upper(), SEVERITY_COLORS['UNKNOWN'])
    cell.fill = PatternFill(
        start_color=color,
        end_color=color,
        fill_type='solid'
    )


# ============================================================================
# Константы для ширины колонок (для разных типов листов)
# ============================================================================

# Ширина колонок для SCA листа
SCA_COLUMN_WIDTHS = {
    'A': 8,   # №
    'B': 40,  # Источник
    'C': 50,  # Путь
    'D': 25,  # Пакет
    'E': 20,  # Версия
    'F': 20,  # Идентификатор уязвимости
    'G': 15,  # Уровень критичности
    'H': 15,  # Срок устранения
    'I': 25,  # Статус
    'J': 40   # Комментарий
}

# Ширина колонок для PTAI листа (9 колонок)
PTAI_COLUMN_WIDTHS = {
    'A': 5,   # №
    'B': 15,  # ID уязвимости
    'C': 40,  # Тип уязвимости
    'D': 50,  # Класс и метод / Уязвимый файл
    'E': 40,  # Комментарий
    'F': 15,  # Статус
    'G': 15,  # CWSS
    'H': 40,  # Дата устранения (формула)
    'I': 40   # Компенсирующие меры
}


# Для обратной совместимости с существующим кодом
HEADER_FONT = HEADER_FONT
HEADER_FILL = HEADER_FILL
CELL_BORDER = THIN_BORDER
INFO_FONT = INFO_FONT